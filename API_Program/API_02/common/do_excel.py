# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from API_Program.API_02.common import project_path
from API_Program.API_02.common.read_config import ReadConfig

class DoExcel:
    '''完成测试数据的读取以及测试结果的写回'''
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name    #Excel工作簿文件名或地址
        self.sheet_name=sheet_name  #表单名

    def read_data(self):
        '''从Excel读取数据，有返回值'''
        #拿到配置文件的case_id
        case_id=ReadConfig(project_path.conf_path).get_data('CASE','case_id')
        wb=load_workbook(self.file_name)
        st=wb[self.sheet_name]
        #开始读取数据，
        test_data=[]
        for i in range(2,st.max_row+1): #每行
            row_data={} # 每行的存到字典中，以每列的title作为key
            row_data['CaseId']=st.cell(i,1).value
            row_data['Module']=st.cell(i,2).value #第一列
            row_data['Title']=st.cell(i,3).value
            row_data['Url']=st.cell(i,4).value
            row_data['Method']=st.cell(i,5).value
            row_data['Params']=st.cell(i,6).value
            row_data['ExpectedResult']=st.cell(i,7).value
            test_data.append(row_data)

        wb.close()
        final_test_data=[]  #存最终的用例数据
        if case_id == 1:    #等于1，执行所有用例
            final_test_data=test_data
        # 否则，如果是列表，获取列表里的数字执行指定的用例
        else:
            for i in case_id:    #遍历配置文件中case_id的值
                final_test_data.append(test_data[i-1]) #case_id=1,为test_data的第一条用例
        return final_test_data

    def write_back(self,row,col,value):
        '''写回测试结果到Excel中'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        sheet.cell(row,col).value=value
        wb.save(self.file_name)
        wb.close()

if __name__ == '__main__':
    file_name=project_path.case_path
    sheet_name='test_case'
    test_data=DoExcel(file_name,sheet_name).read_data()
    print(test_data)

