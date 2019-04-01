# -*-coding:utf-8-*-
# @Time      :2019/2/24/024 10:41
# @Author    :Tanxi
# @Email     :1410510771@qq.com
# @File      :openpyxl_0224.py
# @Software  :PyCharm Community Edition

# 1：写一个类 类的作用是完成Excel数据的读写 新建表单的操作
# 函数一：读取指定表单的数据，有一个列表row_list,把每一行的每一个单元格的数据存到row_list里面去。
# 每一行都有 一个单独的row_list [[1,2,3],[4,5,6]]
# 每一行数据读取完毕后，把row_list存到大列表all_row_list
# 函数二：在指定的单元格写入指定的数据，并保存到当前Excel
# 函数三：新建一个Excel
from openpyxl import load_workbook,workbook
from week_06.homework.python_logging import MyLogger

log=MyLogger()

class PyExcel:
    '''类的作用：是完成Excel数据的读，写，新建表单的操作'''
    try:
        def read_store(self,file,sheet):
            '''
            将Excel中指定表单中每行数据存为一个列表row_list,再将其作为子列表存入all_row_list中
            :param file: 文件名
            :param sheet: 表单名
            :return: all_row_list--一个嵌套列表
            '''

            log.info('开始读取数据:')
            all_row_list=[]#把每行数据作为子列表，存放在该列表中
            wb=load_workbook(file)#打开excel文件
            sheet=wb[sheet]#定位表单
            for item in range(1,sheet.max_row+1):
                row_list=[]#存放每行单元格内数据，每一行先将其置空
                for i in range(1,sheet.max_column+1):
                    row_list.append(sheet.cell(row=item,column=i).value)
                all_row_list.append(row_list)
            log.info('读取到的数据为：{}'.format(all_row_list))
            log.info('读取数据完毕')
            return all_row_list
    except Exception as e:
        log.error('读取文件错误，请检查文件名或表单名是否正确')
        raise e

    try:
        def write_data(self,file,sheet,row,column,value):
            '''
            往指定Excel文件的指定表单指定行列插入指定数据
            :param file: 文件名
            :param sheet: 表单名
            :param row: 表单行
            :param column: 表单列
            :param value: 要写入的数据
            :return: None
            '''
            log.info('开始写入数据:')
            wb = load_workbook(file)  # 打开excel文件
            sheet = wb[sheet]  # 定位表单
            sheet.cell(row,column).value=value#将值写入指定单元格
            wb.save(file)
            wb.close()#写文件后一定记得关闭
            log.info('写入的数据为：{}'.format(value))
            log.info('写入数据完毕')
    except Exception as e:
        log.error('写入数据出错，请检查文件名及相关信息是否正确')
        raise  e

    try:
        def create_excel(self,file):
            '''
            新建一个Excel文件
            :param file: 文件名
            :return:
            '''
            wb=workbook.Workbook(file)
            wb.save(file)
            log.info('新建一个文件，文件名为：{}'.format(file))
            return wb
    except Exception as e:
        log.error('新建文件出错')
        raise  e

if __name__ == '__main__':
    res=PyExcel().read_store('风花雪月.xlsx','Sheet1')
    print(res)
    PyExcel().create_excel('test.xlsx')
    PyExcel().write_data('test.xlsx','Sheet',3,2,'3行2列测试数据')




# 2：思考：分别将我们学过的数据类型  int float boolean str list tuple dict 写到每个单元格里面，
# 观察，你通过openpyxl操作后拿到的数据分别是是什么类型。

#int float boolean str:通过openpyxl读取到的数据均是该数据类型本身
#list tuple dict：通过openpyxl读取到的数据为字符串（str）类型

# wb=load_workbook('data_type.xlsx')#打开excel文件
# sheet=wb['Sheet']#定位表单
# list=[]
# for item in range(1, sheet.max_column + 1):
#     for i in range(1, sheet.max_row + 1):
#         res=sheet.cell(row=i, column=item).value
#         print(type(res))
#         list.append(res)
# print(list)
