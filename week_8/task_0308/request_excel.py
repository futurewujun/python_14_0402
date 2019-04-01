# -*- coding: utf-8 -*-
# 要求如下： 1）http请求类（可以根据传递的method--get/post完成不同的请求），要求有返回值。
# 2）测试用例的数据存储在Excel中，并编写一个从Excel中读取数据的测试类，包含的函数能够读取测试数据，并且能够写回测试结果，要求有返回值。
# 3）新建一个run.py文件， 在这里面完成Excel数据的读取以及完成用例的执行，并写回测试结果到Excel文档里面。 至此已经完成了接口自动化测试的第一步。

import requests
from openpyxl import load_workbook
class HttpRequest:
    '''根据传递的method--get/post完成不同的请求'''
    def __init__(self,url,param,method):
        self.url=url
        self.param=param
        self.method=method

    def http_request(self):
        if self.method.lower() == 'get':
            res=requests.get(self.url,params=self.param)
        else:
            res=requests.post(self.url,data=self.param)
        return res.text

class DoExcel:
    '''从excel中测试数据，并且能够写回测试结果，要求有返回值'''
    def read_excel(self):
        '''读取数据'''
        wb=load_workbook('APITestCases.xlsx')   #打开文件
        st=wb['register']   #定位表单
        all_row=[]    #存全部的数据
        for i in range(2,st.max_row+1):
            each_row=[] #存每行的数据
            for j in range(1,st.max_column):
                res=st.cell(i,j).value
                each_row.append(res)
            all_row.append(each_row)
        return all_row

    def write_result(self,row,column,value):
        '''写回测试结果'''
        wb=load_workbook('APITestCases.xlsx')
        st=wb['register']
        st.cell(row,column).value=value    #写数据
        wb.save('APITestCases.xlsx')

if __name__ == '__main__':

    doex=DoExcel()
    print(doex.read_excel())
    doex.write_result(row=2,column=8,value='qwe')



