# -*- coding: utf-8 -*-
from openpyxl import load_workbook
class DoExcel:
    '''读取指定表单的数据，以嵌套列表形式存储，每一行都是一个子列表，每一个单元格都是子列表里面的元素'''
    def read_sheet_data(self):
        b=[]    #存表单
        wb=load_workbook('new_py.xlsx')
        st=wb['Sheet1']
        for i in range(1,st.max_row+1):
            a = []  # 存每行
            for j in range(1,st.max_column+1):
                res=st.cell(i,j).value
                if res != None:
                    a.append(res)
                    if a not in b:  #不存在才加
                        b.append(a)
        print(b)
ex=DoExcel()
ex.read_sheet_data()

file_path = r"new_py.xlsx"
import openpyxl
# 打开文件
wb = openpyxl.load_workbook(file_path)
# 读取文件的所有表格
Q = wb.get_sheet_by_name('Sheet1')
# 获取最大行数
row_max = Q.max_row
# 获取最大列
con_max = Q.max_column
# 把上面写入内容打印在控制台
list1 = []  # 最后结果
list2 = []  # 数据缓存放入列表
for i in range(1, row_max + 1):
    for j in range(1, con_max + 1):
        list2.append(Q.cell(row=i, column=j).value)
    list1.append(list2)
    list2 = []
# print(list1)