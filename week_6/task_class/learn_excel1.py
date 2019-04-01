# -*- coding: utf-8 -*-
from openpyxl import load_workbook
# # load_workbook()       # 同时支持读与写
# # 打开工作簿
# # wb=load_workbook('py_14.xlsx')      # 相对目录
# wb=load_workbook('E:\PycharmProjects\python_14\week_6\py_14.xlsx')  # 绝对路径
# # 定位表单
# sheet=wb['Sheet1']
# # sheet=wb.get_sheet_by_name('Sheet')   # 旧方法
# # 定位单元格
# # st=sheet.cell(row=4,column=3).value #   # 第4行第3列
# st=sheet.cell(4,1).value    #   不指定的话，就是默认为 行，列
# # print(st) # 输出
# # 可以得到最大的行数，最大的列数
# print(sheet.max_row)
# print(sheet.max_column)
# # 把第6行的数据全部读出来
# for i in range(1,sheet.max_column+1):
#     res=sheet.cell(6,i).value
#     # print(res)
# # 把所有行，所有列全部都读出来
# # for i in range(1,sheet.max_row+1):
# #     for j in range(1,sheet.max_column+1):
# #         res1=sheet.cell(i,j).value
# #         if res1 != None:    # 去掉None的
#             # print(res1)
# # 最后一行如果有空格，就把后面的全部删除掉
# # 写数据：
# # sheet.cell(3,2,'作者：李白')   # 方法一
# sheet.cell(3,1).value='作者：未来'    # 方法二
'''只要进行了单元格的更改都要去进行保存'''
# # wb.save('py_15.xlsx')   # 将打开的py_14 另存为py_15
# # wb.save('py_14.xlsx')   # 如果保存在打开的文件中，要先将文件手动关闭

# 新建工作簿
from openpyxl import workbook
wb=workbook.Workbook('lemon.xlsx')  # 因为要保存，保存时要写名字，所以这里可以不写名字
wb.save('lemon.xlsx')
# 新建了保存之后要再打开才能进行操作

# 循环写数据，写到不同的表单里面     --- 类与对象的问题








