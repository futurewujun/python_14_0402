# -*- coding: utf-8 -*-
# 1：写一个类 类的作用是完成Excel数据的读写 新建表单的操作
# #函数一：读取指定表单的数据， #有一个列表row_list,把每一行的每一个单元格的数据存到row_list里面去。
#          每一行都有 一个单独的row_list [[1,2,3],[4,5,6]] #每一行数据读取完毕后，
#         把row_list存到大列表all_row_list
# #函数二：在指定的单元格写入指定的数据，并保存到当前Excel
# #函数三：新建一个Excel
from openpyxl import load_workbook
from openpyxl import workbook
class DoExcel:
    '''完成Excel数据的读写 新建表单的操作'''
    def read_date(self,file_name,sheet_name):
        '''读取指定表单的数据，以嵌套列表形式存储，每一行都是一个子列表，每一个单元格都是子列表里面的元素'''
        all_row_list = []  # 存表单的数据
        wb = load_workbook(file_name)  # 打开工作簿
        st = wb[sheet_name]  # 定位表单
        for i in range(1, st.max_row + 1):  # 行数
            row_list = []   # 存每行的数据     # 定义接收每行数据的列表为空，继续循环接收下一行数据
            for j in range(1, st.max_column + 1):  # 列数
                res = st.cell(i, j).value
                if res != None:
                    row_list.append(res)  # 接收每行数据放于列表
            all_row_list.append(row_list)  # 将每行列表再放入总列表
        print(all_row_list)

    def alter_data(self,file_name,row,col,msg):
        '''在指定的单元格写入指定的数据，并保存到当前Excel'''
        wb = load_workbook(file_name)    #打开工作簿
        st = wb['Sheet1']   #定位表单
        st.cell(row,col).value=msg    #第三行第四列，写入数据  --- 写数据之前必须手动关闭文件
        wb.save(file_name)   #保存
        res = st.cell(row,col).value    #定位单元格，读数据
        print(res)

    def new_excel(self,file_name,sheet_name):
        '''新建一个Excel'''
        wb = workbook.Workbook()    #新建excel
        wb.create_sheet(sheet_name)  #新建表单
        wb.save(file_name)

if __name__ == '__main__':
    exa = DoExcel()
    # exa.alter_data('py_14.xlsx',4,3,'明日之后')
    # exa.new_excel('py123.xlsx','pyt0hon')
    exa.read_date('py_14.xlsx','Sheet1')


