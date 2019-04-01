# -*- coding: utf-8 -*-
# 2：思考：分别将我们学过的数据类型  int float boolean str list tuple dict 写到每个单元格里面，
#    观察，你通过openpyxl操作后拿到的数据分别是什么类型。
from openpyxl import load_workbook
from openpyxl import workbook
def input_data(msg):
    wb = workbook.Workbook()    #新建
    wb.save('tk_question1.xlsx')     #保存 ---修改或者新建之后必须先保存
    wb_1 = load_workbook('tk_question1.xlsx')    #打开工作簿
    st = wb_1['Sheet'] #定位表单
    st.cell(row=1,column=1).value=msg  # 修改之前必须手动关闭文件
    wb_1.save('tk_question1.xlsx')
    res = st.cell(1,1).value
    print(res)
input_data(1)
input_data(100.53)
input_data(True)
input_data('字符串')
input_data(['列','表'])
input_data(('元','组'))
input_data({'name':'lemon','age':'18'})
# wb = load_workbook('tk_question.xlsx')
# st = wb['Sheet']
# for i in range(1,st.max_column+1):
#     res = st.cell(1,i).value
#     print(res)




