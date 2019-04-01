# -*- coding: utf-8 -*-
from openpyxl import load_workbook
class Do:
    def read_date(self,file_name,sheet_name):
        '''读取指定表单的数据，以嵌套列表形式存储，每一行都是一个子列表，每一个单元格都是子列表里面的元素'''
        row_list = []  # 存每行的数据
        all_row_list = []   #存表单的数据
        wb = load_workbook(file_name)  # 打开工作簿
        st = wb[sheet_name]  # 定位表单
        '''循环每行的数据存到列表中，再定义列表为空接收下一行数据'''
        for i in range(1,st.max_row+1):  #行数
            for j in range(1,st.max_column+1):  #列数
                res = st.cell(i, j).value
                row_list.append(res)        #接收每行数据放于列表
            all_row_list.append(row_list)    #将每行列表再放入总列表
            row_list = []   #定义接收每行数据的列表为空，继续循环接收下一行数据
        print(all_row_list)
e=Do()
e.read_date('new_py.xlsx','Sheet1')