# -*- coding: utf-8 -*-
# 要求如下： 1）http请求类（可以根据传递的method--get/post完成不同的请求），要求有返回值。
# 2）测试用例的数据存储在Excel中，并编写一个从Excel中读取数据的测试类，包含的函数能够读取测试数据，并且能够写回测试结果，要求有返回值。
# 3）新建一个run.py文件， 在这里面完成Excel数据的读取以及完成用例的执行，并写回测试结果到Excel文档里面。 至此已经完成了接口自动化测试的第一步。

import requests
from openpyxl import load_workbook
from API_Program.Task.API_01.common import project_path
class HttpRequest:
    '''根据传递的method--get/post完成不同的请求'''
    def http_request(self,method,url,param,cookies):
        '''根据请求方法来决定发起get请求还是post请求
                method: get post http的请求方式
                url:发送请求的接口地址
                param:随接口发送的请求参数 以字典格式传递
                rtype:有返回值，返回结果是响应报文
                '''
        if method.upper() == 'GET':
            try:
                res=requests.get(url,params=param,cookies=cookies)
            except Exception as e:
                print('get请求出错了：{}'.format(e))
        elif method.upper() == 'PST':
            try:
                res=requests.post(url,data=param)
            except Exception as e:
                print('post请求出错了：{}'.format(e))
        else:
            print('不支持该种方式')
            res = None
        return res

class DoExcel:
    '''从excel中测试数据，并且能够写回测试结果，要求有返回值'''
    def read_excel(self):
        '''读取数据'''
        wb=load_workbook(project_path.case_path)   #打开文件
        st=wb['register']   #定位表单
        all_row=[]    #存全部的数据
        for i in range(2,st.max_row+1):
            each_row=[] #存每行的数据
            for j in range(1,st.max_column-1):
                res=st.cell(i,j).value
                each_row.append(res)
            all_row.append(each_row)
        wb.close()  #操作了文件都要关闭
        return all_row

    def write_result(self,row,column,value):
        '''写回测试结果'''
        wb=load_workbook(project_path.case_path)
        st=wb['test_case']
        st.cell(row,column).value=value    #写数据
        wb.save(project_path.case_path)
        wb.close()

if __name__ == '__main__':

    # doex=DoExcel()
    # print(doex.read_excel())


    http=HttpRequest()
    method='get'
    url='http://47.107.168.87:8080/futureloan/mvc/api/member/login'
    param={"mobilephone":"18412342586","pwd":"wl123456"}
    res=http.http_request(method,url,param,cookies=None)
    print(res.text)
    print(res.cookies)
    res1=http.http_request('get','http://47.107.168.87:8080/futureloan/mvc/api/loan/getLoanList',cookies=res.cookies,param=None)    #加入cookies，获取loanId
    print(res1.text)
    res2=http.http_request('get','http://47.107.168.87:8080/futureloan/mvc/api/member/lis',cookies=res.cookies,param=None)  #获取memberId   -- 好像不行
    # print(res2.text)
    # 充值
    res3=http.http_request('get','http://47.107.168.87:8080/futureloan/mvc/api/member/recharge',{"mobilephone":"18412342586","amount":"100"},cookies=res.cookies)
    # print(res3.json())


